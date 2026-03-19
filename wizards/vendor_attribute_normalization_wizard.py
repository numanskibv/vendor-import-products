import base64
import io
import re

from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError


class VendorAttributeNormalizationWizard(models.TransientModel):
    _name = "vendor.attribute.normalization.wizard"
    _description = "Normalisatie van attributen voor import"

    vendor_id = fields.Many2one(
        comodel_name="res.partner",
        string="Leverancier",
        domain=[("supplier_rank", ">", 0)],
        help="De leverancier waarvoor deze regels worden opgeslagen.",
    )
    file = fields.Binary(string="Importbestand")
    filename = fields.Char(string="Bestandsnaam")
    attribute_type = fields.Selection(
        [("color", "Kleur"), ("size", "Maat")], string="Attribuut", required=True
    )
    raw_values = fields.Text(string="Gevonden waarden", readonly=True)
    mapping_ids = fields.One2many(
        comodel_name="vendor.attribute.normalization.mapping",
        inverse_name="wizard_id",
        string="Normalisatie mapping",
    )

    def action_extract_values(self):
        """Lees het Excel-bestand, vind de attribuutkolom en extraheer unieke waarden."""
        self.ensure_one()
        if not self.file:
            raise UserError(_("Upload eerst een importbestand."))
        if not self.attribute_type:
            raise UserError(_("Kies eerst een attribuut-type (Kleur of Maat)."))

        try:
            import openpyxl
        except ImportError:
            raise UserError(
                _(
                    'Python-pakket "openpyxl" is vereist. Installeer dit in de Odoo-omgeving.'
                )
            )

        file_content = base64.b64decode(self.file)
        try:
            workbook = openpyxl.load_workbook(
                filename=io.BytesIO(file_content), data_only=True, read_only=True
            )
            worksheet = workbook.worksheets[0]
        except Exception as e:
            raise UserError(
                _("Het geüploade Excel-bestand kan niet worden gelezen: %s") % e
            )

        # Vind de header-rij
        header_row = None
        best_nonempty = 0
        for row in worksheet.iter_rows(min_row=1, max_row=10, values_only=True):
            nonempty = sum(1 for cell in row if cell is not None and str(cell).strip())
            if nonempty > best_nonempty:
                best_nonempty = nonempty
                header_row = list(row)

        if not header_row:
            raise UserError(_("Geen geldige header-rij gevonden in het Excel-bestand."))

        headers_raw = [str(h).strip() if h is not None else "" for h in header_row]

        def _norm_header(value):
            txt = str(value or "").strip().lower()
            txt = re.sub(r"[^a-z0-9]+", "", txt)
            return txt

        # Vind de juiste kolom op basis van attribuut-type
        keywords = {
            "color": ["kleur", "color", "colour", "farbe"],
            "size": ["maat", "size", "grootte", "groesse"],
        }.get(self.attribute_type, [])

        target_col_idx = -1
        for idx, header in enumerate(headers_raw):
            norm_h = _norm_header(header)
            if any(kw in norm_h for kw in keywords):
                target_col_idx = idx
                break

        if target_col_idx == -1:
            raise UserError(
                _(
                    "Kon geen kolom vinden voor attribuut '%s'. Zorg dat de kolomkop een van de volgende trefwoorden bevat: %s"
                )
                % (self.attribute_type, ", ".join(keywords))
            )

        # Extraheer unieke waarden
        unique_values = set()
        for row in worksheet.iter_rows(min_row=len(worksheet[1]) + 1, values_only=True):
            if len(row) > target_col_idx:
                cell_value = row[target_col_idx]
                if cell_value is not None:
                    value_str = str(cell_value).strip()
                    if value_str:
                        unique_values.add(value_str)

        if not unique_values:
            self.raw_values = (
                _("Geen waarden gevonden in kolom '%s'.") % headers_raw[target_col_idx]
            )
            self.mapping_ids = [(5, 0, 0)]  # Verwijder bestaande mappings
            return

        sorted_values = sorted(list(unique_values))

        # Update de wizard velden
        self.raw_values = "\n".join(sorted_values)
        self.mapping_ids = [(5, 0, 0)]  # Verwijder bestaande mappings

        mapping_vals = []
        for val in sorted_values:
            mapping_vals.append((0, 0, {"raw_value": val, "normalized_value": val}))

        self.mapping_ids = mapping_vals

    def action_apply_mapping(self):
        """Sla de mapping-regels op als persistente normalisatieregels."""
        self.ensure_one()
        if not self.mapping_ids:
            raise UserError(_("Er zijn geen mapping-regels om op te slaan."))

        Rule = self.env["vendor.attribute.normalization.rule"]
        created_count = 0
        updated_count = 0

        for mapping in self.mapping_ids:
            raw_value = mapping.raw_value.strip()
            normalized_value = mapping.normalized_value.strip()
            if not raw_value or not normalized_value:
                continue

            # Zoek een bestaande regel
            domain = [
                ("vendor_id", "=", self.vendor_id.id),
                ("attribute_type", "=", self.attribute_type),
                ("raw_value", "=", raw_value),
            ]
            existing_rule = Rule.search(domain, limit=1)

            if existing_rule:
                if existing_rule.normalized_value != normalized_value:
                    existing_rule.normalized_value = normalized_value
                    updated_count += 1
            else:
                Rule.create(
                    {
                        "vendor_id": self.vendor_id.id,
                        "attribute_type": self.attribute_type,
                        "raw_value": raw_value,
                        "normalized_value": normalized_value,
                    }
                )
                created_count += 1

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Normalisatieregels opgeslagen"),
                "message": _("Regels aangemaakt: %s\nRegels bijgewerkt: %s")
                % (created_count, updated_count),
                "type": "success",
                "sticky": False,
            },
        }


class VendorAttributeNormalizationMapping(models.TransientModel):
    _name = "vendor.attribute.normalization.mapping"
    _description = "Mapping van ruwe naar genormaliseerde waarde"

    wizard_id = fields.Many2one(
        comodel_name="vendor.attribute.normalization.wizard", string="Wizard"
    )
    raw_value = fields.Char(string="Ruwe waarde", required=True)
    normalized_value = fields.Char(string="Genormaliseerde waarde", required=True)

    @api.constrains("raw_value", "wizard_id")
    def _check_unique_raw_value(self):
        for record in self:
            if not record.wizard_id:
                continue
            domain = [
                ("wizard_id", "=", record.wizard_id.id),
                ("raw_value", "=", record.raw_value),
                ("id", "!=", record.id),
            ]
            if self.search_count(domain) > 0:
                raise ValidationError(
                    _("De ruwe waarde '%s' bestaat al in deze mapping.")
                    % record.raw_value
                )
