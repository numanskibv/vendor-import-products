import base64
import hashlib
import io
import logging
import re
from os.path import basename

from odoo import api, fields, models, _
from odoo.exceptions import UserError


_logger = logging.getLogger(__name__)


class VendorImportWizard(models.TransientModel):
    _name = "vendor.import.wizard"
    _description = "Vendor Import Wizard"

    wizard_record_id = fields.Integer(
        string="Wizard Record ID",
        compute="_compute_wizard_record_id",
        readonly=True,
    )

    def _compute_wizard_record_id(self):
        for wizard in self:
            wizard.wizard_record_id = wizard.id or 0

    file = fields.Binary(string="File")
    filename = fields.Char(string="Filename")
    vendor_id = fields.Many2one(
        comodel_name="res.partner",
        string="Supplier",
        domain=[("supplier_rank", ">", 0)],
        help="Used to validate the file belongs to the selected supplier (via Supplier Reference) and to create vendor pricelist lines (supplierinfo).",
    )
    manual_profile_id = fields.Many2one(
        comodel_name="vendor.import.profile",
        string="Importprofiel",
        domain="[('active', '=', True)]",
        help="Kies handmatig een importprofiel als het automatisch niet wordt gevonden.",
    )
    vendor_manual_html = fields.Html(
        string="Manual",
        compute="_compute_vendor_manual_html",
        readonly=True,
        help="Shows the Excel column naming manual for the selected supplier.",
    )

    def _default_margin_percentage(self):
        """Return the default margin percentage from system config.

        Falls back to 30.0 if the config parameter is not set or invalid.
        """
        icp = self.env["ir.config_parameter"].sudo()
        value = icp.get_param(
            "vendor_import_module.default_margin_percentage", default="30.0"
        )
        try:
            return float(value)
        except (TypeError, ValueError):
            return 30.0

    margin_percentage = fields.Float(
        string="Margin (%)",
        default=lambda self: self._default_margin_percentage(),
    )
    overwrite_prices = fields.Boolean(string="Overwrite Prices", default=False)
    archive_missing_variants = fields.Boolean(
        string="Archive Missing Variants", default=True
    )

    publish_products = fields.Boolean(
        string="Publish Products on Website",
        default=True,
        help=(
            "When enabled, products with an 'is_published' flag in the Excel "
            "will be published on the website during import. This wizard "
            "never unpublishes products."
        ),
    )

    apply_category_rules = fields.Boolean(
        string="Apply Category Rules",
        default=True,
        help=(
            "When enabled, product templates created during this import "
            "will be matched against vendor category rules to auto-assign "
            "a product category. Existing templates keep their category."
        ),
    )

    create_new_categories = fields.Boolean(
        string="Create New Categories",
        default=False,
        help=(
            "When enabled, an optional category column in the Excel file "
            "(e.g. 'Categorie' or 'Category') will be used to assign a "
            "product category to templates without a category. If no "
            "matching category exists, a new top-level category will be "
            "created automatically."
        ),
    )

    category_creation_scan_header = fields.Char(
        string="Kolomnaam voor categorieën (Excel)",
        help=(
            "Plak hier de Excel-kolomnaam (header) waarvan de waarden gebruikt moeten worden "
            "om productcategorieën te bepalen wanneer 'Create New Categories' aan staat. "
            "Vergelijking is niet hoofdlettergevoelig en negeert extra spaties/tekens."
        ),
    )

    # Legacy fields kept for backward compatibility; no longer shown/used.
    category_creation_scan_column_id = fields.Many2one(
        comodel_name="vendor.import.wizard.column",
        string="Category scan column (legacy)",
        readonly=True,
    )
    excel_column_ids = fields.One2many(
        comodel_name="vendor.import.wizard.column",
        inverse_name="wizard_id",
        string="Excel columns (legacy)",
        readonly=True,
    )

    category_creation_scan_field = fields.Selection(
        selection=[
            ("category_name", "Excel category column"),
            ("brand_name", "Brand column"),
        ],
        string="Category scan field",
        default="category_name",
        help=(
            "Select which field from the Excel should be used to derive product "
            "categories when 'Create New Categories' is enabled. This affects both "
            "the analysis report (missing categories) and the import step (category assignment/creation)."
        ),
    )

    def _extract_excel_headers(self):
        """Return header strings from the uploaded Excel.

        This intentionally keeps validation light; vendor/profile checks happen
        later in the actual parsing.
        """
        self.ensure_one()
        if not self.file:
            return []

        try:
            import openpyxl
        except ImportError as exc:
            _logger.exception("openpyxl is not installed")
            raise UserError(
                _(
                    'Python-pakket "openpyxl" is vereist. Installeer dit in de Odoo-omgeving.'
                )
            ) from exc

        file_content = base64.b64decode(self.file)
        try:
            workbook = openpyxl.load_workbook(
                filename=io.BytesIO(file_content), data_only=True, read_only=True
            )
        except Exception as exc:
            _logger.exception("Failed to read Excel file while extracting headers")
            raise UserError(
                _("Het geüploade Excel-bestand kan niet worden gelezen.")
            ) from exc

        worksheet = workbook.worksheets[0]
        header_row = None
        best_nonempty = 0
        for row in worksheet.iter_rows(min_row=1, max_row=10, values_only=True):
            nonempty = sum(1 for cell in row if cell is not None and str(cell).strip())
            if nonempty > best_nonempty:
                best_nonempty = nonempty
                header_row = list(row)

        if not header_row or best_nonempty == 0:
            return []
        return [str(h).strip() if h is not None else "" for h in header_row]

    def _refresh_excel_columns(self):
        """Persist a cache of Excel headers for this wizard.

        The Many2one dropdown needs DB-backed records; we therefore refresh the
        cache on write/create when the file changes.
        """
        Column = self.env["vendor.import.wizard.column"]

        def _norm_header(value):
            if value is None:
                return ""
            txt = str(value).strip().lower()
            txt = re.sub(r"[^a-z0-9]+", " ", txt)
            txt = re.sub(r"\s+", " ", txt).strip()
            return txt

        for wizard in self:
            if not wizard.id:
                continue

            prev_selected = wizard.category_creation_scan_column_id
            prev_index = prev_selected.index if prev_selected else None
            prev_header = (prev_selected.header or "").strip() if prev_selected else ""

            # Always clear the legacy selection before deleting cached columns.
            wizard.with_context(skip_refresh_excel_columns=True).write(
                {"category_creation_scan_column_id": False}
            )

            Column.search([("wizard_id", "=", wizard.id)]).unlink()

            headers = wizard._extract_excel_headers() if wizard.file else []
            create_vals = []
            for idx, header in enumerate(headers or []):
                create_vals.append(
                    {
                        "wizard_id": wizard.id,
                        "header": header or "",
                        "index": idx,
                    }
                )
            created = Column.create(create_vals) if create_vals else Column.browse()

            # Restore previous legacy selection if possible (by index first, then header).
            restore_col = False
            if created and prev_index is not None:
                restore_col = created.filtered(lambda c: c.index == prev_index)[:1]
            if not restore_col and created and prev_header:
                restore_col = created.filtered(
                    lambda c: (c.header or "").strip() == prev_header
                )[:1]
            if restore_col:
                wizard.with_context(skip_refresh_excel_columns=True).write(
                    {"category_creation_scan_column_id": restore_col.id}
                )
                continue

            # Pick a sensible default for legacy field (first header that looks like a category column).
            if wizard.file and created:
                candidates = []
                for col in created:
                    h = _norm_header(col.header)
                    if h in {
                        "categorie",
                        "category",
                        "kategorie",
                        "product category",
                        "productcategorie",
                        "product categorie",
                    }:
                        candidates.append(col)
                    elif "categ" in h:
                        candidates.append(col)
                default_col = candidates[0] if candidates else False
                if default_col:
                    wizard.with_context(skip_refresh_excel_columns=True).write(
                        {"category_creation_scan_column_id": default_col.id}
                    )

    @api.onchange("file")
    def _onchange_file_refresh_excel_columns(self):
        # In a modal wizard the record is often created immediately; refreshing
        # here makes the dropdown fill right after selecting the file.
        for wizard in self:
            if wizard.file and wizard.id:
                wizard._refresh_excel_columns()

    @api.model_create_multi
    def create(self, vals_list):
        records = super().create(vals_list)
        if self.env.context.get("skip_refresh_excel_columns"):
            return records
        for rec, vals in zip(records, vals_list):
            if vals.get("file"):
                rec._refresh_excel_columns()
        return records

    def write(self, vals):
        res = super().write(vals)
        if self.env.context.get("skip_refresh_excel_columns"):
            return res
        if "file" in vals or "filename" in vals:
            self._refresh_excel_columns()
        return res

    # Analysis / rule-drafting helpers
    analysis_log = fields.Text(
        string="Analysis log",
        readonly=True,
        help=(
            "Report produced by the Analyse step, showing how many templates "
            "match existing category rules and suggestions for new rules."
        ),
    )

    create_draft_rules = fields.Boolean(
        string="Allow creating draft rules",
        default=False,
        help=(
            "When enabled, the 'Create draft rules' action can create "
            "vendor.category.rule records based on analysis suggestions."
        ),
    )

    draft_vendor_code = fields.Selection(
        selection=lambda self: self.env[
            "vendor.category.rule"
        ]._selection_vendor_code(),
        string="Draft vendor code",
        help=(
            "Vendor code to use on newly created draft category rules. "
            "Defaults to the current supplier's code (e.g. tgh) when available."
        ),
    )

    draft_priority_base = fields.Integer(
        string="Base priority for drafts",
        default=100,
        help=(
            "Base priority used for auto-created draft rules. A small offset "
            "based on keyword length will be added on top of this value."
        ),
    )

    draft_active = fields.Boolean(
        string="Draft rules active",
        default=False,
        help=(
            "Whether auto-created draft rules should be active immediately. "
            "By default they are created inactive so they can be reviewed first."
        ),
    )

    test_passed = fields.Boolean(string="Test passed", readonly=True)
    test_file_sha1 = fields.Char(string="Test file sha1", readonly=True)
    test_report_html = fields.Html(string="Test report", readonly=True)

    image_queue_pending_count = fields.Integer(
        string="Images pending",
        compute="_compute_image_queue_counts",
        readonly=True,
    )
    image_queue_done_count = fields.Integer(
        string="Images done",
        compute="_compute_image_queue_counts",
        readonly=True,
    )
    image_queue_error_count = fields.Integer(
        string="Images error",
        compute="_compute_image_queue_counts",
        readonly=True,
    )
    image_queue_total_count = fields.Integer(
        string="Images total",
        compute="_compute_image_queue_counts",
        readonly=True,
    )

    @api.depends("vendor_id")
    def _compute_image_queue_counts(self):
        queue = self.env["vendor.import.image.queue"].sudo()
        for wizard in self:
            if not wizard.vendor_id:
                wizard.image_queue_pending_count = 0
                wizard.image_queue_done_count = 0
                wizard.image_queue_error_count = 0
                wizard.image_queue_total_count = 0
                continue

            vendor_domain = [("vendor_id", "=", wizard.vendor_id.id)]
            pending = queue.search_count(vendor_domain + [("state", "=", "pending")])
            done = queue.search_count(vendor_domain + [("state", "=", "done")])
            error = queue.search_count(vendor_domain + [("state", "=", "error")])
            wizard.image_queue_pending_count = pending
            wizard.image_queue_done_count = done
            wizard.image_queue_error_count = error
            wizard.image_queue_total_count = pending + done + error

    def action_process_images_now(self):
        self.ensure_one()
        if not self.vendor_id:
            raise UserError(_("Selecteer eerst een leverancier."))

        result = (
            self.env["vendor.import.image.queue"]
            .sudo()
            ._process_pending_jobs(domain=[("vendor_id", "=", self.vendor_id.id)])
        )
        pending = (
            self.env["vendor.import.image.queue"]
            .sudo()
            .search_count(
                [("vendor_id", "=", self.vendor_id.id), ("state", "=", "pending")]
            )
        )
        message = _(
            "Afbeeldingenwachtrij verwerkt voor %(vendor)s. Gereed: %(done)s, Mislukt: %(failed)s, Nog in wachtrij: %(pending)s"
        ) % {
            "vendor": self.vendor_id.display_name,
            "done": result.get("done", 0),
            "failed": result.get("failed", 0),
            "pending": pending,
        }

        # Refresh the queue counters on the wizard so the UI reflects
        # the latest state after processing.
        self._compute_image_queue_counts()

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Afbeeldingenwachtrij"),
                "message": message,
                "type": "info",
                "sticky": False,
            },
        }

    def action_reset_image_queue(self):
        """Delete all image queue jobs for the selected supplier.

        This allows starting with a clean queue before running a new import
        or image processing run.
        """
        self.ensure_one()
        if not self.vendor_id:
            raise UserError(_("Selecteer eerst een leverancier."))

        queue = self.env["vendor.import.image.queue"].sudo()
        domain = [("vendor_id", "=", self.vendor_id.id)]
        jobs = queue.search(domain)
        count = len(jobs)
        if count:
            jobs.unlink()

        message = _(
            "Afbeeldingenwachtrij leeggemaakt voor %(vendor)s. Verwijderde taken: %(count)s"
        ) % {
            "vendor": self.vendor_id.display_name,
            "count": count,
        }

        # Refresh the queue counters on the wizard so the UI immediately
        # shows that the queue has been cleared.
        self._compute_image_queue_counts()

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Afbeeldingenwachtrij"),
                "message": message,
                "type": "info",
                "sticky": False,
            },
        }

    @api.onchange("file", "vendor_id", "margin_percentage")
    def _onchange_reset_test(self):
        for wizard in self:
            wizard.test_passed = False
            wizard.test_file_sha1 = False
            wizard.test_report_html = False

    def _get_supplier_prefix(self, vendor):
        """Derive supplier prefix from reference, with fallback to name."""
        if not vendor:
            return ""
        prefix = (vendor.ref or "").strip().upper()
        if not prefix and "TGH" in (vendor.display_name or "").upper():
            return "TGH"
        return prefix

    @api.onchange("vendor_id")
    def _onchange_vendor_id_set_draft_vendor_code(self):
        for wizard in self:
            supplier_prefix = self._get_supplier_prefix(wizard.vendor_id)
            code = (supplier_prefix or "any").lower()
            wizard.draft_vendor_code = code

    def _parse_tgh_excel(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Upload eerst een bestand om te importeren."))
        if not self.vendor_id:
            raise UserError(_("Selecteer eerst een leverancier."))

        # Supplier Reference is used as the filename prefix safeguard.
        # For the TGH supplier we can safely fall back to 'TGH' when the name contains it,
        # because we also validate the file content with a TGH fingerprint check below.
        supplier_prefix = (self.vendor_id.ref or "").strip().upper()
        if not supplier_prefix:
            vendor_name_u = (self.vendor_id.display_name or "").upper()
            if "TGH" in vendor_name_u:
                supplier_prefix = "TGH"
                _logger.warning(
                    "Supplier '%s' has no Reference; assuming prefix 'TGH' based on name.",
                    self.vendor_id.display_name,
                )

        if not supplier_prefix:
            raise UserError(
                _(
                    "De geselecteerde leverancier heeft geen Referentie (Interne referentie).\n\n"
                    "Waarom dit belangrijk is: deze import controleert of het geüploade bestand bij de leverancier hoort aan de hand van het bestandsnaamprefix.\n\n"
                    "Oplossing: open de leverancier '%(supplier)s' en vul de Referentie in, bijvoorbeeld 'TGH', en probeer de import daarna opnieuw."
                )
                % {"supplier": self.vendor_id.display_name}
            )

        filename_raw = (self.filename or "").strip()
        filename = basename(filename_raw) if filename_raw else ""
        if filename:
            if not filename.startswith(supplier_prefix):
                raise UserError(
                    _(
                        "Verkeerd bestand voor de geselecteerde leverancier.\n\n"
                        "Verwacht: bestandsnaam moet beginnen met de leveranciersreferentie '%(prefix)s' (hoofdletters).\n"
                        "Voorbeeld: '%(prefix)s producten ....xlsx'\n\n"
                        "Geselecteerde leverancier: %(supplier)s\n"
                        "Leveranciersreferentie: %(prefix)s\n"
                        "Geselecteerd bestand: %(filename)s\n\n"
                        "Oplossing: kies de juiste leverancier in de wizard of hernoem het bestand zodat het begint met '%(prefix)s'."
                    )
                    % {
                        "prefix": supplier_prefix,
                        "supplier": self.vendor_id.display_name,
                        "filename": filename,
                    }
                )
        else:
            _logger.warning(
                "Missing filename for vendor.import.wizard %s (supplier=%s). Proceeding with content fingerprint validation.",
                self.id,
                self.vendor_id.display_name,
            )

        try:
            import openpyxl
        except ImportError as exc:
            _logger.exception("openpyxl is not installed")
            raise UserError(
                _(
                    'Python-pakket "openpyxl" is vereist. Installeer dit in de Odoo-omgeving.'
                )
            ) from exc

        def _norm_header(value):
            if value is None:
                return ""
            txt = str(value).strip().lower()
            txt = re.sub(r"[^a-z0-9]+", " ", txt)
            txt = re.sub(r"\s+", " ", txt).strip()
            return txt

        def _cell_to_text(value):
            if value is None:
                return ""
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value).strip()

        def _cell_to_float(value):
            if value is None or value == "":
                return None
            if isinstance(value, (int, float)):
                return float(value)
            txt = str(value).strip()
            if not txt:
                return None
            txt = txt.replace("€", "").replace(" ", "").replace(",", ".")
            return float(txt)

        def _unique_urls(values):
            seen = set()
            out = []
            for v in values or []:
                if not v:
                    continue
                if v in seen:
                    continue
                seen.add(v)
                out.append(v)
            return out

        file_content = base64.b64decode(self.file)
        file_sha1 = hashlib.sha1(file_content).hexdigest()
        try:
            workbook = openpyxl.load_workbook(
                filename=io.BytesIO(file_content), data_only=True
            )
        except Exception as exc:
            _logger.exception("Failed to read Excel file")
            raise UserError(
                _("Het geüploade Excel-bestand kan niet worden gelezen.")
            ) from exc

        worksheet = workbook.worksheets[0]

        header_row = None
        best_nonempty = 0
        for row in worksheet.iter_rows(min_row=1, max_row=10, values_only=True):
            nonempty = sum(1 for cell in row if cell is not None and str(cell).strip())
            if nonempty > best_nonempty:
                best_nonempty = nonempty
                header_row = list(row)

        if not header_row or best_nonempty == 0:
            raise UserError(_("Het geüploade Excel-bestand is leeg."))

        headers_raw = [str(h).strip() if h is not None else "" for h in header_row]
        header_to_index = {str(h).strip(): idx for idx, h in enumerate(headers_raw)}

        # Probeer eerst een dynamisch importprofiel (vendor.import.profile)
        # te vinden op basis van leverancier/bestandsnaam/headers.
        dynamic_profile = None
        dynamic_col_index = {}
        dynamic_image_indices = []
        try:
            dynamic_profile, dynamic_col_index, dynamic_image_indices = (
                self.env["vendor.import.profile"]
                .sudo()
                .match_profile(
                    vendor=self.vendor_id,
                    filename=filename,
                    headers_raw=headers_raw,
                    norm_func=_norm_header,
                )
            )
        except Exception as exc:
            _logger.exception("Error while matching vendor.import.profile: %s", exc)

        # Als geen automatisch profiel gevonden, probeer handmatig geselecteerd profiel
        if not dynamic_profile and self.manual_profile_id:
            manual_profile = self.manual_profile_id
            if manual_profile.vendor_id and manual_profile.vendor_id != self.vendor_id:
                raise UserError(
                    _("Het geselecteerde profiel hoort bij een andere leverancier.")
                )

            # Controleer of het profiel matcht met de headers
            role_to_index = {}
            image_indices = []
            for line in manual_profile.line_ids:
                pattern_norm = _norm_header(line.header_pattern or "")
                found_idx = None
                for idx, h in enumerate(headers_raw):
                    h_norm = _norm_header(h)
                    if h_norm.startswith(pattern_norm) or pattern_norm in h_norm:
                        found_idx = idx
                        break
                if found_idx is not None:
                    role_to_index[line.role] = found_idx
                    if (line.role or "").startswith("image_"):
                        if found_idx not in image_indices:
                            image_indices.append(found_idx)
                elif line.required:
                    raise UserError(
                        _(
                            "Vereiste kolom '%(role)s' niet gevonden in headers: %(pattern)s"
                        )
                        % {"role": line.role, "pattern": line.header_pattern}
                    )

            dynamic_profile = manual_profile
            dynamic_col_index = role_to_index
            dynamic_image_indices = image_indices

        # Vendor fingerprint checks (helps prevent mixing vendor files)
        # Alleen toepassen als we geen dynamisch profiel hebben gevonden en
        # het om TGH gaat.
        if not dynamic_profile and supplier_prefix == "TGH":
            link_header = "Link naar TGH.nl"
            link_idx = None
            for idx, h in enumerate(headers_raw):
                if str(h).strip() == link_header:
                    link_idx = idx
                    break

            looks_like_tgh = False
            if link_idx is not None:
                for row in worksheet.iter_rows(min_row=2, max_row=60, values_only=True):
                    if not row or link_idx >= len(row):
                        continue
                    cell = row[link_idx]
                    if cell and "tgh.nl" in str(cell).lower():
                        looks_like_tgh = True
                        break

            if not looks_like_tgh and any(
                "tghimages" in str(h).lower() for h in headers_raw
            ):
                looks_like_tgh = True

            if not looks_like_tgh:
                raise UserError(
                    _(
                        "This file does not look like a TGH export (no TGH links found in '%(col)s').\n"
                        "Selected supplier: %(supplier)s\n"
                        "Detected headers: %(headers)s"
                    )
                    % {
                        "col": link_header,
                        "supplier": self.vendor_id.display_name,
                        "headers": ", ".join([h for h in headers_raw if h]),
                    }
                )

        # Bepaal kolomindices voor de kernrollen (product_name, color, size,
        # purchase_price, sku, barcode). Als er een dynamisch profiel is,
        # gebruiken we dat; anders vallen we terug op de oude TGH-profielen.

        image_col_indices = list(dynamic_image_indices or [])

        if dynamic_profile:
            chosen_profile = dynamic_profile.code or dynamic_profile.name or "dynamic"
            col_index = dict(dynamic_col_index or {})
        else:
            raise UserError(
                _(
                    "Geen passend importprofiel gevonden voor leverancier %(vendor)s.\n"
                    "Maak een profiel aan in Voorraad → Vendor Import → Importprofielen."
                )
                % {"vendor": self.vendor_id.display_name}
            )

        # Hard-validate core role mappings to prevent KeyError crashes later.
        required_roles = ["product_name", "color", "size", "purchase_price", "sku"]
        missing_roles = [r for r in required_roles if col_index.get(r) is None]
        if missing_roles:

            def _suggest_headers_for_role(role):
                keywords_by_role = {
                    "product_name": [
                        "product",
                        "artikel",
                        "omschrijving",
                        "naam",
                        "name",
                        "title",
                        "description",
                    ],
                    "color": ["kleur", "color", "colour"],
                    "size": [
                        "maat",
                        "size",
                        "afmet",
                        "dimension",
                        "lengte",
                        "breedte",
                        "hoogte",
                    ],
                    "purchase_price": [
                        "inkoop",
                        "purchase",
                        "cost",
                        "net",
                        "prijs",
                        "price",
                        "excl",
                        "eur",
                    ],
                    "sku": ["sku", "artikel", "code", "item", "productcode", "artnr"],
                }
                kws = keywords_by_role.get(role, [])
                scored = []
                for idx, header in enumerate(headers_raw or []):
                    h_norm = _norm_header(header)
                    score = sum(1 for k in kws if k in h_norm)
                    if score:
                        scored.append((score, idx, header))
                scored.sort(key=lambda t: (-t[0], t[1]))
                return [f"kolom {idx + 1}: {header}" for _, idx, header in scored[:5]]

            mapping_lines = []
            for role, idx in sorted((col_index or {}).items()):
                if idx is None:
                    continue
                if isinstance(idx, int) and 0 <= idx < len(headers_raw):
                    mapping_lines.append(
                        f"- {role} -> kolom {idx + 1}: {headers_raw[idx]}"
                    )
                else:
                    mapping_lines.append(f"- {role} -> kolom {idx}")
            mapping_hint = (
                "\n".join(mapping_lines) if mapping_lines else "(geen mapping gevonden)"
            )

            headers_lines = []
            for idx, header in enumerate(headers_raw or []):
                if header:
                    headers_lines.append(f"- kolom {idx + 1}: {header}")
            headers_hint = (
                "\n".join(headers_lines[:60])
                if headers_lines
                else "(geen headers gevonden)"
            )
            if len(headers_lines) > 60:
                headers_hint += "\n- ..."

            suggestions_lines = []
            for role in missing_roles:
                sugg = _suggest_headers_for_role(role)
                if sugg:
                    suggestions_lines.append(f"{role}: " + "; ".join(sugg))
            suggestions_hint = (
                "\n".join(suggestions_lines)
                if suggestions_lines
                else "(geen suggesties)"
            )

            raise UserError(
                _(
                    "Importprofiel '%(profile)s' mist vereiste rollen/kolommen: %(missing)s\n\n"
                    "Oplossing: controleer het importprofiel (Voorraad → Vendor Import → Importprofielen) en zorg dat deze rollen gemapt zijn: "
                    "product_name, color, size, purchase_price, sku.\n\n"
                    "Gedetecteerde headers:\n%(headers)s\n\n"
                    "Suggesties (mogelijke kolommen per rol):\n%(suggestions)s\n\n"
                    "Huidige mapping:\n%(mapping)s"
                )
                % {
                    "profile": chosen_profile,
                    "missing": ", ".join(missing_roles),
                    "headers": headers_hint,
                    "suggestions": suggestions_hint,
                    "mapping": mapping_hint,
                }
            )

        # Optional publish column (e.g. is_published / published / publish / online)
        # We normalize the header (similar to _norm_header) so that variants like
        # "Is Published" or "is published" also work.
        is_published_idx = None
        publish_header_candidates = {
            "is_published",
            "is published",
            "published",
            "publish",
            "online",
            "website",
            "web shop",
            "webshop",
            "shop",
            "publiceer",
            "publiceren",
            "zichtbaar",
            "visible",
        }
        for raw_header, idx in header_to_index.items():
            hnorm = _norm_header(raw_header)
            if hnorm in publish_header_candidates:
                is_published_idx = idx
                break

        # Optional brand column (e.g. Merk / Brand)
        brand_idx = None
        # Optional category column (e.g. Categorie / Category)
        category_idx = None

        # Collect non-fatal issues to report in Test/Analyse.
        warnings = []

        # Prefer explicit profile mapping when available.
        if col_index.get("brand") is not None:
            brand_idx = col_index.get("brand")
        if col_index.get("category_name") is not None:
            category_idx = col_index.get("category_name")

        # Optional override: user-provided Excel header name for category creation.
        # When set, we treat that column as the category source (both analysis
        # and import will see it as category_name).
        scan_header = (self.category_creation_scan_header or "").strip()
        if scan_header:
            wanted = _norm_header(scan_header)
            found_idx = None
            for idx, raw in enumerate(headers_raw):
                if _norm_header(raw) == wanted:
                    found_idx = idx
                    break
            if found_idx is not None:
                category_idx = found_idx
            else:
                warnings.append(
                    "Category scan column not found in Excel headers: %s" % scan_header
                )

        for raw_header, idx in header_to_index.items():
            hnorm = _norm_header(raw_header)
            if brand_idx is None and hnorm in {"merk", "brand"}:
                brand_idx = idx
            elif category_idx is None and hnorm in {"categorie", "category"}:
                category_idx = idx

        products_data = {}
        errors = []
        header_len = len(header_row)
        rows_detected = 0

        # Barcode is optional in some vendor sheets / profiles.
        barcode_idx = col_index.get("barcode")
        if barcode_idx is None:
            barcode_header_candidates = {
                "barcode",
                "bar code",
                "ean",
                "ean13",
                "ean-13",
                "gtin",
                "gtin13",
                "gtin-13",
                "streepjescode",
            }
            for raw_header, idx in header_to_index.items():
                hnorm = _norm_header(raw_header)
                if hnorm in barcode_header_candidates:
                    barcode_idx = idx
                    col_index["barcode"] = idx
                    break
        if barcode_idx is None:
            warnings.append("No barcode column detected; barcodes will be left empty.")

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not row or not any(
                cell is not None and str(cell).strip() for cell in row
            ):
                continue
            rows_detected += 1
            row = list(row) + [None] * max(0, header_len - len(row))

            product_name = _cell_to_text(row[col_index["product_name"]])
            if not product_name:
                continue

            color = _cell_to_text(row[col_index["color"]])
            size = _cell_to_text(row[col_index["size"]])
            purchase_price = _cell_to_float(row[col_index["purchase_price"]])
            sku = _cell_to_text(row[col_index["sku"]])
            barcode = ""
            if barcode_idx is not None and barcode_idx < len(row):
                barcode = _cell_to_text(row[barcode_idx])

            brand_name = ""
            if brand_idx is not None and brand_idx < len(row):
                brand_name = _cell_to_text(row[brand_idx])

            category_name = ""
            if category_idx is not None and category_idx < len(row):
                category_name = _cell_to_text(row[category_idx])

            is_published = None
            if is_published_idx is not None and is_published_idx < len(row):
                flag_text = _cell_to_text(row[is_published_idx]).strip().lower()
                if flag_text in {
                    "1",
                    "true",
                    "yes",
                    "y",
                    "ja",
                    "published",
                    "online",
                    "x",
                }:
                    is_published = True
                elif flag_text in {"0", "false", "no", "n", "nee"}:
                    is_published = False

            image_urls = []
            for img_idx in image_col_indices:
                url = _cell_to_text(row[img_idx])
                if url:
                    image_urls.append(url)
            image_urls = _unique_urls(image_urls)

            if not color or not size:
                errors.append(
                    f"Row {row_number}: Color and Size are required (product={product_name})."
                )
                continue

            if purchase_price is None:
                purchase_price = 0.0
                warnings.append(
                    f"Row {row_number}: missing purchase_price; defaulting to 0.0 (product={product_name}, sku={sku or ''})"
                )

            template_bucket = products_data.setdefault(
                product_name,
                {
                    "rows": [],
                    "colors": set(),
                    "sizes": set(),
                    "sale_prices": set(),
                    "image_urls": set(),
                },
            )
            template_bucket["rows"].append(
                {
                    "row_number": row_number,
                    "color": color,
                    "size": size,
                    "purchase_price": purchase_price,
                    "sku": sku,
                    "barcode": barcode,
                    "brand_name": brand_name,
                    "category_name": category_name,
                    "is_published": is_published,
                    "image_urls": image_urls,
                }
            )
            template_bucket["colors"].add(color)
            template_bucket["sizes"].add(size)
            sale_price = purchase_price * (
                1.0 + (self.margin_percentage or 0.0) / 100.0
            )
            template_bucket["sale_prices"].add(round(sale_price, 6))
            for url in image_urls:
                template_bucket["image_urls"].add(url)

        if not products_data:
            errors.append("No data rows found to import.")

        meta = {
            "supplier_prefix": supplier_prefix,
            "filename": filename,
            "file_sha1": file_sha1,
            "chosen_profile": chosen_profile,
            "rows_detected": rows_detected,
            "templates_count": len(products_data),
            "image_columns": len(image_col_indices),
        }

        return products_data, meta, errors, warnings

    # ============================================================
    # Analysis helpers (category coverage & draft rule suggestions)
    # ============================================================

    def _suggest_keyword_for_name(self, product_name):
        """Suggest a keyword (and optional category path) for an uncategorized name.

        Heuristics:
        - First look for known tokens like 'softshell', 'poloshirt', ...
        - If none found, fall back to top non-trivial word in the name,
          excluding common stopwords.
        """

        name_raw = (product_name or "").strip()
        if not name_raw:
            return None

        norm = name_raw.replace("’", "'").lower()

        token_map = {
            "softshell": "Bedrijfskleding / Jassen / Softshell",
            "fleece": "Bedrijfskleding / Jassen / Fleece",
            "bodywarmer": "Bedrijfskleding / Bodywarmers",
            "jacket": "Bedrijfskleding / Jassen",
            "windjacket": "Bedrijfskleding / Jassen",
            "3-in-1": "Bedrijfskleding / Jassen",
            "hoodie": "Bedrijfskleding / Sweaters & Hoodies",
            "hooded": "Bedrijfskleding / Sweaters & Hoodies",
            "sweater": "Bedrijfskleding / Sweaters & Hoodies",
            "zipneck": "Bedrijfskleding / Vesten",
            "sweatjacket": "Bedrijfskleding / Vesten",
            "poloshirt": "Bedrijfskleding / Polo's",
            "polo": "Bedrijfskleding / Polo's",
            "t-shirt": "Bedrijfskleding / T-Shirts",
            "beanie": "Bedrijfskleding / Accessoires",
            "apron": "Bedrijfskleding / Accessoires",
        }

        # Known token based suggestion
        for token, category_path in token_map.items():
            if token in norm:
                return {"keyword": token, "category_path": category_path}

        # Fallback: pick a non-trivial word as keyword
        words = re.findall(r"[a-z0-9]+", norm)
        stopwords = {
            "mens",
            "women",
            "unisex",
            "kids",
            "heren",
            "dames",
            "kind",
            "kinder",
            "tgh",
            "the",
            "and",
            "voor",
            "met",
        }
        candidates = [w for w in words if w not in stopwords and len(w) >= 3]
        if not candidates:
            return None

        keyword = candidates[0]
        return {"keyword": keyword, "category_path": None}

    def _collect_category_analysis(self):
        """Collect category coverage and keyword suggestions for the current file.

        Returns (products_data, meta, matched_names, unmatched_info, suggestions_by_keyword).
        unmatched_info maps product name -> suggestion dict or None.
        suggestions_by_keyword maps keyword -> {"category_path", "examples"}.
        """

        products_data, meta, errors, warnings = self._parse_tgh_excel()
        if errors:
            raise UserError("\n".join(errors[:30]))

        CategoryRule = self.env["vendor.category.rule"]
        vendor_code = (meta.get("supplier_prefix") or "").lower() or "any"

        unique_names = sorted(products_data.keys())
        matched_names = set()
        unmatched_info = {}
        suggestions_by_keyword = {}

        for name in unique_names:
            rule = CategoryRule.match(name, vendor_code)
            if rule:
                matched_names.add(name)
                continue

            suggestion = self._suggest_keyword_for_name(name)
            unmatched_info[name] = suggestion
            if suggestion and suggestion.get("keyword"):
                kw = suggestion["keyword"]
                entry = suggestions_by_keyword.setdefault(
                    kw,
                    {
                        "category_path": suggestion.get("category_path"),
                        "examples": set(),
                    },
                )
                entry["examples"].add(name)

        return (
            products_data,
            meta,
            matched_names,
            unmatched_info,
            suggestions_by_keyword,
        )

    def action_test_tgh(self):
        self.ensure_one()

        try:
            # Ensure product variants feature is enabled, since the import
            # relies on color/size-based variants.
            if not self.env.user.has_group("product.group_product_variant"):
                raise UserError(
                    _(
                        "Productvarianten zijn uitgeschakeld in de voorraadinstellingen.\n\n"
                        "Deze import is afhankelijk van varianten (kleur/maat). Schakel "
                        "'Productvarianten' in onder Voorraad \u2192 Configuratie \u2192 Instellingen "
                        "en voer de test daarna opnieuw uit."
                    )
                )

            products_data, meta, errors, warnings = self._parse_tgh_excel()
        except UserError as exc:
            self.write(
                {
                    "test_passed": False,
                    "test_file_sha1": False,
                    "test_report_html": f"<h3>Test failed</h3><pre>{str(exc)}</pre>",
                }
            )
            return {
                "type": "ir.actions.client",
                "tag": "display_notification",
                "params": {
                    "title": _("TGH Test"),
                    "message": str(exc),
                    "type": "danger",
                    "sticky": True,
                },
            }

        templates = len(products_data)
        rows = sum(len(b.get("rows") or []) for b in products_data.values())
        combos = 0
        duplicate_combos = 0
        unique_urls = set()
        sku_seen = {}
        ean_seen = {}
        for product_name, bucket in products_data.items():
            combo_set = set()
            for row in bucket.get("rows") or []:
                key = (row.get("color") or "", row.get("size") or "")
                if key in combo_set:
                    duplicate_combos += 1
                combo_set.add(key)
                combos += 1

                sku = (row.get("sku") or "").strip()
                if sku:
                    sku_seen.setdefault(sku, []).append(product_name)
                ean = (row.get("barcode") or "").strip()
                if ean:
                    ean_seen.setdefault(ean, []).append(product_name)

                for url in row.get("image_urls") or []:
                    unique_urls.add(url)
            for url in bucket.get("image_urls") or []:
                unique_urls.add(url)

        dup_skus = [k for k, v in sku_seen.items() if len(v) > 1]
        dup_eans = [k for k, v in ean_seen.items() if len(v) > 1]

        passed = len(errors) == 0

        report_parts = [
            "<h3>TGH-importtest</h3>",
            "<ul>",
            f"<li><b>Profiel</b>: {meta.get('chosen_profile')}</li>",
            f"<li><b>Rijen</b>: {rows} (gedetecteerd: {meta.get('rows_detected')})</li>",
            f"<li><b>Templates</b>: {templates}</li>",
            f"<li><b>Varianten (rijen/combos)</b>: {combos}</li>",
            f"<li><b>Unieke afbeelding-URL's</b>: {len(unique_urls)}</li>",
            "</ul>",
        ]

        if errors:
            report_parts.append("<h4>Fouten</h4><ul>")
            for msg in errors[:30]:
                report_parts.append(f"<li>{msg}</li>")
            report_parts.append("</ul>")

        if warnings:
            report_parts.append("<h4>Waarschuwingen</h4><ul>")
            for msg in warnings[:30]:
                report_parts.append(f"<li>{msg}</li>")
            report_parts.append("</ul>")

        if duplicate_combos:
            report_parts.append(
                f"<p><b>Warning</b>: duplicate color/size rows found: {duplicate_combos}</p>"
            )
            report_parts.append(
                f"<p><b>Waarschuwing</b>: dubbele kleur/maat-rijen gevonden: {duplicate_combos}</p>"
            )
        if dup_skus:
            report_parts.append(
                f"<p><b>Warning</b>: duplicate SKUs detected (showing up to 10): {', '.join(dup_skus[:10])}</p>"
            )
            report_parts.append(
                f"<p><b>Waarschuwing</b>: dubbele SKU's gedetecteerd (maximaal 10 getoond): {', '.join(dup_skus[:10])}</p>"
            )
        if dup_eans:
            report_parts.append(
                f"<p><b>Warning</b>: duplicate EANs detected (showing up to 10): {', '.join(dup_eans[:10])}</p>"
            )
            report_parts.append(
                f"<p><b>Waarschuwing</b>: dubbele EAN's gedetecteerd (maximaal 10 getoond): {', '.join(dup_eans[:10])}</p>"
            )

        if passed:
            report_parts.append(
                "<p><b>OK</b>: Test passed. You can run Import now.</p>"
            )
            report_parts.append(
                "<p><b>OK</b>: Test geslaagd. Je kunt nu Import uitvoeren.</p>"
            )
        else:
            report_parts.append(
                "<p><b>Not OK</b>: Fix the errors above and re-run the test.</p>"
            )
            report_parts.append(
                "<p><b>Niet OK</b>: Los de bovenstaande fouten op en voer de test opnieuw uit.</p>"
            )

        self.write(
            {
                "test_passed": passed,
                "test_file_sha1": meta.get("file_sha1"),
                "test_report_html": "".join(report_parts),
            }
        )

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Test Structuur bestand"),
                "message": _("Test geslaagd.") if passed else _("Test mislukt."),
                "type": "success" if passed else "danger",
                "sticky": True,
            },
        }

    def action_analyse_import_file(self):
        """Analyse the Excel for category rule coverage and suggestions.

        Does not modify products; only fills analysis_log and shows a notification.
        """
        self.ensure_one()

        try:
            products_data, meta, matched_names, unmatched_info, suggestions = (
                self._collect_category_analysis()
            )
        except UserError as exc:
            self.analysis_log = str(exc)
            return {
                "type": "ir.actions.client",
                "tag": "display_notification",
                "params": {
                    "title": _("Analyse"),
                    "message": str(exc),
                    "type": "danger",
                    "sticky": True,
                },
            }

        total_templates = len(products_data)
        matched_count = len(matched_names)
        unmatched_count = total_templates - matched_count
        vendor_code = (meta.get("supplier_prefix") or "").lower() or "any"

        lines = []
        lines.append(f"{vendor_code.upper()} categorie-analyse")
        lines.append("=====================")
        lines.append(f"Totaal aantal templates: {total_templates}")
        lines.append(f"Gematcht door regels: {matched_count}")
        lines.append(f"Geen match: {unmatched_count}")
        lines.append("")

        if unmatched_count:
            lines.append("Templates zonder match (eerste 50):")
            for name in sorted(unmatched_info.keys())[:50]:
                suggestion = unmatched_info.get(name)
                if suggestion and suggestion.get("keyword"):
                    kw = suggestion["keyword"]
                    cat = suggestion.get("category_path") or "(geen categorievoorstel)"
                    lines.append(f"- {name} -> zoekwoord: '{kw}' | categorie: {cat}")
                else:
                    lines.append(f"- {name}")

        if suggestions:
            lines.append("")
            lines.append("Voorgestelde zoekwoorden (geaggregeerd):")
            for kw, info in sorted(suggestions.items()):
                cat = info.get("category_path") or "(geen categorievoorstel)"
                example_list = sorted(info.get("examples") or [])[:5]
                examples_str = ", ".join(example_list)
                lines.append(
                    f"- '{kw}' -> categorie: {cat} | voorbeelden: {examples_str}"
                )

        # Category creation analysis (based on selected scan column/field)
        scan_header = (self.category_creation_scan_header or "").strip()
        scan_field = self.category_creation_scan_field or "category_name"
        if scan_header:
            scan_label = f"Excel kolom: {scan_header}"
        else:
            scan_label = {
                "category_name": "Excel categorie-kolom",
                "brand_name": "Merk-kolom",
            }.get(scan_field, scan_field)

        categories_in_file = set()
        for product_name, bucket in products_data.items():
            value = ""
            if scan_header:
                for row in bucket.get("rows") or []:
                    v = (row.get("category_name") or "").strip()
                    if v:
                        value = v
                        break
            elif scan_field == "brand_name":
                for row in bucket.get("rows") or []:
                    v = (row.get("brand_name") or "").strip()
                    if v:
                        value = v
                        break
            else:
                for row in bucket.get("rows") or []:
                    v = (row.get("category_name") or "").strip()
                    if v:
                        value = v
                        break

            if value:
                categories_in_file.add(value)

        if categories_in_file:
            ProductCategory = self.env["product.category"]
            missing = []
            for name in sorted(categories_in_file):
                if not ProductCategory.search([("name", "=ilike", name)], limit=1):
                    missing.append(name)

            lines.append("")
            lines.append("Categorieën uit Excel")
            lines.append("-------------------")
            lines.append(f"Scan veld: {scan_label}")
            lines.append(f"Unieke waarden in bestand: {len(categories_in_file)}")
            lines.append(f"Ontbrekende categorieën in Odoo: {len(missing)}")
            if missing:
                hint = (
                    "Deze worden tijdens import aangemaakt (Create New Categories = aan)."
                    if self.create_new_categories
                    else "Let op: Create New Categories staat uit; deze worden niet automatisch aangemaakt."
                )
                lines.append(hint)
                lines.append("Eerste 50 ontbrekende categorieën:")
                for name in missing[:50]:
                    lines.append(f"- {name}")

        self.analysis_log = "\n".join(lines)

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Analyse"),
                "message": _("Analyse voltooid. Zie het tabblad Analyse voor details."),
                "type": "info",
                "sticky": False,
            },
        }

    def action_create_draft_rules(self):
        """Create draft vendor.category.rule records based on analysis suggestions."""
        self.ensure_one()

        if not self.create_draft_rules:
            raise UserError(
                _(
                    "Schakel eerst 'Conceptregels mogen worden aangemaakt' in "
                    "voordat je deze actie uitvoert, zodat duidelijk is dat je "
                    "concept-categorieregels wilt laten aanmaken."
                )
            )

        products_data, meta, matched_names, unmatched_info, suggestions = (
            self._collect_category_analysis()
        )

        if not suggestions:
            raise UserError(
                _(
                    "Er zijn geen zoekwoordvoorstellen uit de analyse gekomen. "
                    "Voer eerst de stap Analyse uit en controleer of er templates "
                    "zonder match zijn."
                )
            )

        Category = self.env["product.category"]

        # Find or create placeholder category "Te categoriseren" under "Bedrijfskleding".
        root = Category.search(
            [("name", "=", "Bedrijfskleding"), ("parent_id", "=", False)],
            limit=1,
        )
        if not root:
            root = Category.create({"name": "Bedrijfskleding"})

        placeholder = Category.search(
            [
                ("name", "=", "Te categoriseren"),
                ("parent_id", "=", root.id),
            ],
            limit=1,
        )
        if not placeholder:
            placeholder = Category.create(
                {"name": "Te categoriseren", "parent_id": root.id}
            )

        CategoryRule = self.env["vendor.category.rule"]
        vendor_code = self.draft_vendor_code or (
            (meta.get("supplier_prefix") or "").lower() or "any"
        )
        if vendor_code not in {
            c[0] for c in self.env["vendor.category.rule"]._selection_vendor_code()
        }:
            vendor_code = "any"

        base_priority = self.draft_priority_base or 100
        active = self.draft_active

        created = 0
        skipped = 0
        for kw, info in suggestions.items():
            keyword = (kw or "").strip()
            if not keyword:
                continue

            # Skip if an identical rule already exists.
            # Important: include inactive rules (drafts are often created inactive).
            existing = CategoryRule.with_context(active_test=False).search(
                [
                    ("vendor_code", "=", vendor_code),
                    ("keyword", "=", keyword),
                    ("category_id", "=", placeholder.id),
                    ("audience", "=", "any"),
                ],
                limit=1,
            )
            if existing:
                skipped += 1
                continue

            offset = min(50, len(kw))
            vals = {
                "name": f"AUTO: {keyword}",
                "active": active,
                "vendor_code": vendor_code,
                "match_type": "contains",
                "keyword": keyword,
                "priority": base_priority + offset,
                "category_id": placeholder.id,
                "audience": "any",
            }
            try:
                CategoryRule.create(vals)
                created += 1
            except Exception:
                # Race-condition / unexpected duplicate: swallow and continue.
                # The SQL constraint will prevent data corruption.
                skipped += 1

        notice = _("Draft rules created: %(count)s (skipped existing: %(skipped)s)") % {
            "count": created,
            "skipped": skipped,
        }

        if created:
            self.analysis_log = (self.analysis_log or "") + "\n\n" + notice

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Analysis"),
                "message": notice,
                "type": "success" if created else "info",
                "sticky": False,
            },
        }

    def _compute_vendor_manual_html(self):
        for wizard in self:
            vendor = wizard.vendor_id
            if vendor and vendor.vendor_import_excel_manual_html:
                wizard.vendor_manual_html = vendor.vendor_import_excel_manual_html
                continue

            vendor_ref = (vendor.ref or "").strip().upper() if vendor else ""
            if vendor_ref == "TGH":
                wizard.vendor_manual_html = (
                    "<h3>TGH import – kolomnamen</h3>"
                    "<p>Gebruik een TGH-export. De import herkent twee varianten (strict/test export). "
                    "Kolomnamen zijn hoofdlettergevoelig bij de strict TGH-kolommen.</p>"
                    "<h4>Minimaal vereist (test export)</h4>"
                    "<ul>"
                    "<li><b>product_name</b> – productnaam</li>"
                    "<li><b>kleur</b> of <b>color</b> – kleur</li>"
                    "<li><b>maat</b> of <b>size</b> – maat</li>"
                    "<li><b>purchase_price</b> – inkoopprijs</li>"
                    "<li><b>sku</b> – interne referentie / SKU</li>"
                    "<li><b>barcode</b> – EAN (optioneel)</li>"
                    "</ul>"
                    "<h4>Minimaal vereist (strict TGH)</h4>"
                    "<ul>"
                    "<li><b>Artikelnaam</b></li>"
                    "<li><b>Kleur</b></li>"
                    "<li><b>Maat</b></li>"
                    "<li><b>Inkoopprijs</b></li>"
                    "<li><b>SKU</b></li>"
                    "<li><b>Barcode</b> (optioneel)</li>"
                    "</ul>"
                    "<h4>Afbeeldingen (optioneel)</h4>"
                    "<p>Kolommen zoals <b>image0_filename</b>…<b>image6_filename</b> met een URL worden geïmporteerd.</p>"
                    "<h4>Veiligheidschecks</h4>"
                    "<ul>"
                    "<li>Bestandsnaam moet starten met de leveranciersreferentie (bijv. <b>TGH</b>).</li>"
                    "<li>Voor TGH verwacht de import een kolom <b>Link naar TGH.nl</b> met tenminste één tgh.nl link (of headers met tghimages).</li>"
                    "</ul>"
                )
            else:
                wizard.vendor_manual_html = (
                    "<h3>Vendor import – handleiding</h3>"
                    "<p>Vul op de leverancier een handleiding in (Vendor Import Manual). "
                    "Die wordt hier automatisch getoond zodra je een supplier kiest.</p>"
                )

    def action_import_tgh(self):
        self.ensure_one()

        if not self.test_passed:
            raise UserError(
                _(
                    "Voer eerst de Test uit. Importeren is pas mogelijk nadat de test is geslaagd."
                )
            )

        if not self.test_file_sha1:
            raise UserError(_("Teststatus ontbreekt. Voer de Test opnieuw uit."))

        current_sha1 = hashlib.sha1(base64.b64decode(self.file or b""))
        if current_sha1.hexdigest() != self.test_file_sha1:
            raise UserError(
                _(
                    "Het geüploade bestand is gewijzigd sinds de laatste geslaagde test.\n"
                    "Voer de Test opnieuw uit voordat je gaat importeren."
                )
            )

        _logger.info(
            "TGH import started (wizard=%s, vendor=%s, filename=%s)",
            self.id,
            self.vendor_id.id if self.vendor_id else None,
            self.filename,
        )

        products_data, meta, parse_errors, parse_warnings = self._parse_tgh_excel()
        if parse_errors:
            raise UserError("\n".join(parse_errors[:30]))

        ProductTemplate = self.env["product.template"]
        ProductAttribute = self.env["product.attribute"]
        ProductAttributeValue = self.env["product.attribute.value"]
        ImageQueue = self.env["vendor.import.image.queue"]
        CategoryRule = self.env["vendor.category.rule"]
        ProductBrand = self.env["product.brand"]
        ProductCategory = self.env["product.category"]

        try:
            SupplierInfo = self.env["product.supplierinfo"]
        except KeyError:
            SupplierInfo = None
        if self.vendor_id and SupplierInfo is None:
            raise UserError(
                _(
                    "Koppelen aan leveranciers vereist de Inkoop-app (model product.supplierinfo niet gevonden)."
                )
            )

        # Determine vendor code for category rules (e.g. "tgh")
        vendor_code = (meta.get("supplier_prefix") or "").lower() or "any"

        def _ensure_attribute(name):
            attribute = ProductAttribute.search([("name", "=", name)], limit=1)
            if attribute:
                return attribute
            vals = {"name": name}
            if "create_variant" in ProductAttribute._fields:
                vals["create_variant"] = "always"
            return ProductAttribute.create(vals)

        attr_color = _ensure_attribute("Kleur")
        attr_size = _ensure_attribute("Maat")

        value_cache = {}

        def _ensure_value(attribute, value_name):
            key = (attribute.id, value_name)
            if key in value_cache:
                return value_cache[key]
            rec = ProductAttributeValue.search(
                [("attribute_id", "=", attribute.id), ("name", "=", value_name)],
                limit=1,
            )
            if not rec:
                rec = ProductAttributeValue.create(
                    {"attribute_id": attribute.id, "name": value_name}
                )
            value_cache[key] = rec
            return rec

        def _ensure_attribute_line(template, attribute, values):
            line = template.attribute_line_ids.filtered(
                lambda l: l.attribute_id == attribute
            )[:1]
            value_ids = values.ids
            if not line:
                template.write(
                    {
                        "attribute_line_ids": [
                            (
                                0,
                                0,
                                {
                                    "attribute_id": attribute.id,
                                    "value_ids": [(6, 0, value_ids)],
                                },
                            )
                        ]
                    }
                )
                return
            existing = set(line.value_ids.ids)
            missing = [vid for vid in value_ids if vid not in existing]
            if missing:
                line.write({"value_ids": [(4, vid) for vid in missing]})

        products_created = 0
        variants_created = 0
        variants_updated = 0
        variants_archived = 0

        templates_categorized = 0
        templates_uncategorized = 0
        uncategorized_names = set()

        categories_created = 0
        templates_categorized_from_excel = 0

        for product_name, bucket in products_data.items():
            template = ProductTemplate.search([("name", "=", product_name)], limit=1)
            template_created = False
            if not template:
                create_vals = {"name": product_name}
                if "sale_ok" in ProductTemplate._fields:
                    create_vals["sale_ok"] = True
                if "purchase_ok" in ProductTemplate._fields:
                    create_vals["purchase_ok"] = True
                template = ProductTemplate.create(create_vals)
                products_created += 1
                template_created = True

            # Apply category rules on newly created templates (or those without a category)
            if self.apply_category_rules:
                rule = CategoryRule.match(product_name, vendor_code)
                if rule and rule.category_id:
                    # Only override the category for new templates or templates without a category.
                    if template_created or not template.categ_id:
                        _logger.info(
                            "Category rule matched for '%s' (vendor_code=%s): rule=%s, keyword=%s -> category=%s",
                            product_name,
                            vendor_code,
                            rule.display_name,
                            rule.keyword,
                            rule.category_id.display_name,
                        )
                        template.write({"categ_id": rule.category_id.id})
                        templates_categorized += 1
                    else:
                        templates_uncategorized += 1
                        if len(uncategorized_names) < 50:
                            uncategorized_names.add(product_name)
                        _logger.info(
                            "Category rule found for '%s' but existing category kept (vendor_code=%s, current_categ=%s, rule=%s)",
                            product_name,
                            vendor_code,
                            template.categ_id.display_name,
                            rule.display_name,
                        )
                else:
                    templates_uncategorized += 1
                    if len(uncategorized_names) < 50:
                        uncategorized_names.add(product_name)
                    _logger.info(
                        "No category rule match for '%s' (vendor_code=%s)",
                        product_name,
                        vendor_code,
                    )

            # Determine brand for this template (first non-empty brand in its rows)
            template_brand_name = ""
            for row in bucket.get("rows") or []:
                bn = (row.get("brand_name") or "").strip()
                if bn:
                    template_brand_name = bn
                    break

            if template_brand_name:
                # Case-insensitive lookup to avoid duplicate brands with different casing
                brand = ProductBrand.search(
                    [("name", "ilike", template_brand_name)], limit=1
                )
                if not brand:
                    brand = ProductBrand.create({"name": template_brand_name})

                if "brand_id" in template._fields:
                    current_brand = template.brand_id
                    # Do not overwrite existing brand unless overwrite flag is True,
                    # or the template is newly created.
                    if template_created or not current_brand or self.overwrite_prices:
                        template.write({"brand_id": brand.id})

            # Excel-driven category: optional, and only when template has no category yet
            template_category_name = ""
            if (self.category_creation_scan_header or "").strip():
                for row in bucket.get("rows") or []:
                    cn = (row.get("category_name") or "").strip()
                    if cn:
                        template_category_name = cn
                        break
            else:
                scan_field = self.category_creation_scan_field or "category_name"
                if scan_field == "brand_name":
                    for row in bucket.get("rows") or []:
                        cn = (row.get("brand_name") or "").strip()
                        if cn:
                            template_category_name = cn
                            break
                else:
                    for row in bucket.get("rows") or []:
                        cn = (row.get("category_name") or "").strip()
                        if cn:
                            template_category_name = cn
                            break

            if template_category_name and "categ_id" in template._fields:
                if not template.categ_id:
                    category = ProductCategory.search(
                        [("name", "=ilike", template_category_name)], limit=1
                    )
                    if not category and self.create_new_categories:
                        category = ProductCategory.create(
                            {"name": template_category_name}
                        )
                        categories_created += 1

                    if category:
                        template.write({"categ_id": category.id})
                        templates_categorized_from_excel += 1

            icp = self.env["ir.config_parameter"].sudo()
            template_image_only = icp.get_param(
                "vendor_import_module.template_image_only", default="False"
            ).lower() in {"1", "true", "yes", "on"}

            raw_image_urls = [u for u in (bucket.get("image_urls") or []) if u]
            if template_image_only and raw_image_urls:
                # Only keep the first image URL per template to save space
                image_urls_sorted = [sorted(raw_image_urls)[0]]
            else:
                image_urls_sorted = sorted(raw_image_urls)

            template_all = template.with_context(active_test=False)
            before_variant_ids = set(template_all.product_variant_ids.ids)

            color_values = self.env["product.attribute.value"]
            for color_name in sorted(bucket["colors"]):
                color_values |= _ensure_value(attr_color, color_name)

            size_values = self.env["product.attribute.value"]
            for size_name in sorted(bucket["sizes"]):
                size_values |= _ensure_value(attr_size, size_name)

            _ensure_attribute_line(template, attr_color, color_values)
            _ensure_attribute_line(template, attr_size, size_values)

            if hasattr(template, "_create_variant_ids"):
                template._create_variant_ids()

            # Images are processed asynchronously by cron. Enqueue template-level URLs.
            if image_urls_sorted:
                for url in image_urls_sorted:
                    domain = [
                        ("vendor_id", "=", self.vendor_id.id),
                        ("product_tmpl_id", "=", template.id),
                        ("product_id", "=", False),
                        ("url", "=", url),
                    ]
                    if not ImageQueue.search(domain, limit=1):
                        ImageQueue.create(
                            {
                                "vendor_id": self.vendor_id.id,
                                "product_tmpl_id": template.id,
                                "product_id": False,
                                "url": url,
                                "state": "pending",
                            }
                        )

            template_all.invalidate_recordset(["product_variant_ids"])
            after_variant_ids = set(template_all.product_variant_ids.ids)
            variants_created += len(after_variant_ids - before_variant_ids)

            variants_all = template_all.product_variant_ids
            combo_to_variant = {}
            for variant in variants_all:
                found_color = ""
                found_size = ""
                for ptav in variant.product_template_attribute_value_ids:
                    if ptav.attribute_id == attr_color:
                        found_color = (
                            ptav.product_attribute_value_id.name
                            if hasattr(ptav, "product_attribute_value_id")
                            and ptav.product_attribute_value_id
                            else (ptav.name or "")
                        )
                    elif ptav.attribute_id == attr_size:
                        found_size = (
                            ptav.product_attribute_value_id.name
                            if hasattr(ptav, "product_attribute_value_id")
                            and ptav.product_attribute_value_id
                            else (ptav.name or "")
                        )
                if found_color and found_size:
                    combo_to_variant[(found_color, found_size)] = variant

            imported_variant_ids = set()

            for row_data in bucket["rows"]:
                key = (row_data["color"], row_data["size"])
                variant = combo_to_variant.get(key)
                if not variant:
                    _logger.warning(
                        "No variant found for %s (%s/%s) at row %s",
                        product_name,
                        row_data["color"],
                        row_data["size"],
                        row_data["row_number"],
                    )
                    continue

                write_vals = {}

                is_published = row_data.get("is_published")

                if row_data["sku"]:
                    write_vals["default_code"] = row_data["sku"]

                if row_data["barcode"]:
                    write_vals["barcode"] = row_data["barcode"]

                # Variant images: optionally skip per-variant URLs to save space.
                # When template_image_only is enabled, we do not enqueue
                # variant-specific image jobs at all.
                if not template_image_only:
                    # Only enqueue URLs that are not already scheduled at template
                    # level, to avoid duplicate jobs when all varianten dezelfde
                    # foto's delen.
                    row_images = [u for u in (row_data.get("image_urls") or []) if u]
                    variant_only_images = [
                        u for u in row_images if u not in image_urls_sorted
                    ]
                    if variant_only_images:
                        for url in variant_only_images:
                            domain = [
                                ("vendor_id", "=", self.vendor_id.id),
                                ("product_tmpl_id", "=", template.id),
                                ("product_id", "=", variant.id),
                                ("url", "=", url),
                            ]
                            if not ImageQueue.search(domain, limit=1):
                                ImageQueue.create(
                                    {
                                        "vendor_id": self.vendor_id.id,
                                        "product_tmpl_id": template.id,
                                        "product_id": variant.id,
                                        "url": url,
                                        "state": "pending",
                                    }
                                )

                purchase_price = row_data["purchase_price"]
                sale_price = purchase_price * (
                    1.0 + (self.margin_percentage or 0.0) / 100.0
                )

                # Link supplier (vendor) to product via supplierinfo
                if self.vendor_id and SupplierInfo is not None:
                    supplier_domain = [
                        ("partner_id", "=", self.vendor_id.id),
                        ("product_tmpl_id", "=", template.id),
                    ]
                    if "product_id" in SupplierInfo._fields:
                        supplier_domain.append(("product_id", "=", variant.id))

                    supplierinfo = SupplierInfo.search(supplier_domain, limit=1)

                    supplier_vals = {
                        "partner_id": self.vendor_id.id,
                        "product_tmpl_id": template.id,
                    }
                    if "product_id" in SupplierInfo._fields:
                        supplier_vals["product_id"] = variant.id
                    if "min_qty" in SupplierInfo._fields:
                        supplier_vals["min_qty"] = 1.0
                    if "price" in SupplierInfo._fields:
                        if (
                            self.overwrite_prices
                            or not supplierinfo
                            or (supplierinfo.price or 0.0) == 0.0
                        ):
                            supplier_vals["price"] = purchase_price
                    if (
                        "currency_id" in SupplierInfo._fields
                        and self.env.company.currency_id
                    ):
                        supplier_vals["currency_id"] = self.env.company.currency_id.id

                    if supplierinfo:
                        # Update only the fields we set above
                        supplierinfo.write(supplier_vals)
                    else:
                        SupplierInfo.create(supplier_vals)

                if "standard_price" in variant._fields:
                    current_cost = variant.standard_price or 0.0
                    if self.overwrite_prices or current_cost == 0.0:
                        write_vals["standard_price"] = purchase_price

                wrote_sale_price = False
                if (
                    "list_price" in variant._fields
                    and not variant._fields["list_price"].readonly
                ):
                    current_sale = variant.list_price or 0.0
                    if self.overwrite_prices or current_sale == 0.0:
                        write_vals["list_price"] = sale_price
                        wrote_sale_price = True

                if not wrote_sale_price and "list_price" in template._fields:
                    # Fallback: product.template has list_price in standard Odoo.
                    # If imported prices differ per variant, keep the maximum and log it.
                    desired_prices = bucket["sale_prices"]
                    if desired_prices:
                        if len(desired_prices) > 1:
                            _logger.warning(
                                'Multiple sale prices found for "%s"; applying max to template list_price.',
                                product_name,
                            )
                        desired = max(desired_prices)
                        current_template_sale = template.list_price or 0.0
                        if (
                            self.overwrite_prices
                            or template_created
                            or current_template_sale == 0.0
                        ):
                            template.write({"list_price": desired})

                # Optional website publication flag: only ever publish, never unpublish
                if self.publish_products and is_published is True:
                    # Odoo has used both field names across versions/modules.
                    publish_fields = ["is_published", "website_published"]

                    # Try to publish the specific variant if the field exists.
                    for field_name in publish_fields:
                        if field_name in variant._fields and not getattr(
                            variant, field_name, False
                        ):
                            write_vals[field_name] = True

                    # Always ensure the template itself is published if the field exists.
                    for field_name in publish_fields:
                        if field_name in template._fields and not getattr(
                            template, field_name, False
                        ):
                            template.write({field_name: True})
                            break

                if not variant.active:
                    write_vals["active"] = True

                if write_vals:
                    variant.write(write_vals)
                    variants_updated += 1

                imported_variant_ids.add(variant.id)

            if self.archive_missing_variants:
                active_variants = template.product_variant_ids
                to_archive = active_variants.filtered(
                    lambda v: v.id not in imported_variant_ids
                )
                if to_archive:
                    to_archive.write({"active": False})
                    variants_archived += len(to_archive)
                    _logger.info(
                        'Archived %s missing variants for "%s"',
                        len(to_archive),
                        product_name,
                    )

            _logger.info(
                'Imported TGH: "%s" (rows=%s, variants=%s)',
                product_name,
                len(bucket["rows"]),
                len(template_all.product_variant_ids),
            )

        message = _(
            "TGH import finished.\n"
            "Products created: %(products)s\n"
            "Variants created: %(variants_created)s\n"
            "Variants updated: %(variants_updated)s\n"
            "Variants archived: %(variants_archived)s"
        ) % {
            "products": products_created,
            "variants_created": variants_created,
            "variants_updated": variants_updated,
            "variants_archived": variants_archived,
        }

        if self.apply_category_rules:
            category_summary = _(
                "\nTemplates categorized by rules: %(categorized)s\n"
                "Templates without category match or kept existing: %(uncategorized)s"
            ) % {
                "categorized": templates_categorized,
                "uncategorized": templates_uncategorized,
            }

            if uncategorized_names:
                preview = ", ".join(sorted(uncategorized_names)[:20])
                category_summary += _(
                    "\nUncategorized examples (max 20): %(names)s"
                ) % {"names": preview}

            message = f"{message}{category_summary}"

        if self.create_new_categories:
            extra = _(
                "\nCategories created from Excel: %(created)s\n"
                "Templates categorized from Excel category column: %(templ)s"
            ) % {
                "created": categories_created,
                "templ": templates_categorized_from_excel,
            }
            message = f"{message}{extra}"

        _logger.warning("Products created: %s", products_created)
        _logger.warning("Variants created: %s", variants_created)
        _logger.warning("Variants updated: %s", variants_updated)
        _logger.warning("Variants archived: %s", variants_archived)
        _logger.info(message.replace("\n", " | "))

        _logger.info(
            "TGH import finished (wizard=%s, vendor=%s, filename=%s)",
            self.id,
            self.vendor_id.id if self.vendor_id else None,
            self.filename,
        )

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("TGH Import"),
                "message": message,
                "type": "success",
                "sticky": True,
            },
        }

    def open_attribute_normalization_wizard(self):
        self.ensure_one()
        if not self.file:
            raise UserError(
                _("Upload eerst een bestand voordat je de normalisatie-wizard opent.")
            )
        return {
            "type": "ir.actions.act_window",
            "name": _("Normalisatie van attributen"),
            "res_model": "vendor.attribute.normalization.wizard",
            "view_mode": "form",
            "target": "new",
            "context": {
                "default_file": self.file,
                "default_filename": self.filename,
                "default_vendor_id": self.vendor_id.id,
            },
        }
